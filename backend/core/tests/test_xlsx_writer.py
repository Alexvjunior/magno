import io
from datetime import date

from openpyxl import load_workbook

from domain.models import Desocupacao
from domain.xlsx_writer import build_xlsx


def _record(year=2025, month=7, day=3) -> Desocupacao:
    return Desocupacao(
        id="abc",
        status="ACTIVE",
        id_imovel="FLORIANOPOLIS|TOP VISION RESIDENCE|1227",
        cidade="Florianopolis",
        edificio="Top Vision Residence",
        numero_apto="1227",
        area_privativa=68.78,
        tipologia="2 dormitorios",
        uso="Residencial",
        status_evento="Desocupacao",
        data_evento=date(year, month, day),
        data_inicio_contrato=date(2023, 10, 24),
        valor_aluguel=2500.50,
        dias_vacancia=12,
        motivo_desocupacao="Mudou de estado",
        mes=month,
        ano=year,
        criado_por="user-1",
        criado_em="2025-07-03T12:00:00Z",
    )


def test_creates_one_sheet_per_month():
    items = [
        _record(2025, 6, 15),
        _record(2025, 7, 3),
        _record(2025, 7, 20),
    ]
    data = build_xlsx(items)
    wb = load_workbook(io.BytesIO(data))
    assert "Junho 25" in wb.sheetnames
    assert "Julho 25" in wb.sheetnames
    assert wb["Julho 25"].max_row == 5


def test_empty_input_produces_valid_workbook():
    data = build_xlsx([])
    wb = load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) >= 1


def test_header_row_matches_current_form_fields():
    data = build_xlsx([_record()])
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Julho 25"]
    headers = [ws.cell(row=3, column=c).value for c in range(2, 16)]
    assert headers == [
        "Cidade",
        "Edificio",
        "Numero Apto",
        "Area Privativa",
        "Tipologia",
        "Uso",
        "Status Evento",
        "Data Evento",
        "Data Inicio Contrato",
        "Valor Aluguel",
        "Dias Vacancia",
        "Mes",
        "Ano",
        "Motivo Desocupacao",
    ]


def test_writes_current_form_values():
    data = build_xlsx([_record()])
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Julho 25"]
    assert ws.cell(row=4, column=2).value == "Florianopolis"
    assert ws.cell(row=4, column=3).value == "Top Vision Residence"
    assert ws.cell(row=4, column=4).value == "1227"
    assert ws.cell(row=4, column=11).value == 2500.50
