"""Generates a multi-sheet XLSX from Desocupacao records.

One sheet is created per (ano, mes), matching the current frontend form fields.
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Desocupacao

MONTHS_PT = [
    "JANEIRO",
    "FEVEREIRO",
    "MARCO",
    "ABRIL",
    "MAIO",
    "JUNHO",
    "JULHO",
    "AGOSTO",
    "SETEMBRO",
    "OUTUBRO",
    "NOVEMBRO",
    "DEZEMBRO",
]

HEADERS = [
    "Cidade",
    "Edificio",
    "Numero Apto",
    "Area Privativa",
    "Tipologia",
    "Uso",
    "Status Evento",
    "Data Evento",
    "Data Inicio Contrato",
    "Valor Aluguel",
    "Dias Vacancia",
    "Mes",
    "Ano",
    "Motivo Desocupacao",
]


def _sheet_name(d: date) -> str:
    return f"{MONTHS_PT[d.month - 1].title()} {str(d.year)[-2:]}"


def _title(d: date) -> str:
    return f"DESOCUPACOES - {MONTHS_PT[d.month - 1]} {str(d.year)[-2:]}"


def build_xlsx(items: list[Desocupacao]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    by_month: dict[tuple[int, int], list[Desocupacao]] = defaultdict(list)
    for it in items:
        by_month[(it.ano, it.mes)].append(it)

    if not by_month:
        ws = wb.create_sheet("Vazio")
        ws["A1"] = "Sem registros."
        return _to_bytes(wb)

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    center = Alignment(horizontal="center", vertical="center")

    for (year, month) in sorted(by_month.keys()):
        sample_date = date(year, month, 1)
        ws = wb.create_sheet(_sheet_name(sample_date))

        ws.cell(row=2, column=2, value=_title(sample_date)).font = title_font
        ws.cell(row=2, column=2).alignment = center
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + len(HEADERS) - 1)

        for idx, header in enumerate(HEADERS):
            c = ws.cell(row=3, column=2 + idx, value=header)
            c.font = bold
            c.fill = header_fill
            c.alignment = center

        rows = sorted(by_month[(year, month)], key=lambda x: x.data_evento)
        for r_offset, item in enumerate(rows, start=4):
            ws.cell(row=r_offset, column=2, value=item.cidade)
            ws.cell(row=r_offset, column=3, value=item.edificio)
            ws.cell(row=r_offset, column=4, value=item.numero_apto)
            ws.cell(row=r_offset, column=5, value=item.area_privativa).number_format = "0.00"
            ws.cell(row=r_offset, column=6, value=item.tipologia)
            ws.cell(row=r_offset, column=7, value=item.uso)
            ws.cell(row=r_offset, column=8, value=item.status_evento)
            c_evento = ws.cell(row=r_offset, column=9, value=item.data_evento)
            c_evento.number_format = "DD/MM/YYYY"
            c_inicio = ws.cell(row=r_offset, column=10, value=item.data_inicio_contrato)
            c_inicio.number_format = "DD/MM/YYYY"
            ws.cell(row=r_offset, column=11, value=item.valor_aluguel).number_format = "0.00"
            ws.cell(row=r_offset, column=12, value=item.dias_vacancia)
            ws.cell(row=r_offset, column=13, value=item.mes)
            ws.cell(row=r_offset, column=14, value=item.ano)
            ws.cell(row=r_offset, column=15, value=item.motivo_desocupacao)

        widths = [18, 26, 14, 16, 14, 14, 18, 14, 22, 16, 14, 8, 8, 60]
        for idx, width in enumerate(widths):
            ws.column_dimensions[get_column_letter(2 + idx)].width = width

    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
